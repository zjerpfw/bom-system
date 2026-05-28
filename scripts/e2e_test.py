import asyncio
import csv
import json
import os
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

import numpy as np
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine


PROJECT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = PROJECT_DIR / "backend"
sys.path.insert(0, str(BACKEND_DIR))

from app.core.database import Base  # noqa: E402
from app.models.bom_item import BomItem  # noqa: E402
from app.models.material import Material  # noqa: E402
from app.models.missing_material import MissingMaterial  # noqa: E402
from app.services import match_service  # noqa: E402
from app.services.export_service import export_bom_to_excel  # noqa: E402
from app.services.material_service import (  # noqa: E402
    async_import_materials_from_csv,
    async_build_embedding_index,
    load_index,
)


PRODUCT_NAME = "主控板V2"
VECTOR_DIMENSION = 8


SAMPLE_MATERIALS = [
    ("MAT-0001", "铜柱", "M3x6", "个", "五金件"),
    ("MAT-0002", "主控芯片", "STM32F103", "个", "电子元件"),
    ("MAT-0003", "电容100uF", "贴片", "个", "电子元件"),
    ("MAT-0004", "电阻10K", "0603", "个", "电子元件"),
    ("MAT-0005", "电阻1K", "0603", "个", "电子元件"),
    ("MAT-0006", "LED灯", "红色0603", "个", "电子元件"),
    ("MAT-0007", "排针", "2.54mm 2x10", "个", "连接器"),
    ("MAT-0008", "排母", "2.54mm 2x10", "个", "连接器"),
    ("MAT-0009", "USB接口", "Type-C", "个", "连接器"),
    ("MAT-0010", "晶振", "8MHz", "个", "电子元件"),
    ("MAT-0011", "二极管", "SS14", "个", "电子元件"),
    ("MAT-0012", "三极管", "S8050", "个", "电子元件"),
    ("MAT-0013", "保险丝", "1A", "个", "保护器件"),
    ("MAT-0014", "蜂鸣器", "有源5V", "个", "声学器件"),
    ("MAT-0015", "按键", "贴片4脚", "个", "结构件"),
    ("MAT-0016", "端子", "2P 5.08", "个", "连接器"),
    ("MAT-0017", "电感", "10uH", "个", "电子元件"),
    ("MAT-0018", "稳压芯片", "AMS1117-3.3", "个", "电子元件"),
    ("MAT-0019", "PCB板", "主控板V2", "片", "板件"),
    ("MAT-0020", "螺钉", "M3x8", "个", "五金件"),
]


EXTRACTED_BOM = {
    "product": PRODUCT_NAME,
    "items": [
        {"name": "铜柱", "spec": "M3×6", "quantity": 4, "unit": "个"},
        {"name": "主控芯片", "spec": "STM32F103", "quantity": 1, "unit": "个"},
        {"name": "电容100uF", "spec": "贴片", "quantity": 10, "unit": "个"},
        {"name": "未知零件XYZ", "spec": None, "quantity": 1, "unit": "个"},
    ],
}


def deterministic_embedding(texts: list[str]) -> list[list[float]]:
    """生成确定性向量，避免端到端测试消耗OpenAI额度。"""
    vectors = []
    for text in texts:
        seed = sum(ord(char) for char in text)
        vector = [0.0] * VECTOR_DIMENSION
        vector[seed % VECTOR_DIMENSION] = 1.0
        vector[(seed // VECTOR_DIMENSION) % VECTOR_DIMENSION] += 0.25
        vectors.append(vector)
    return vectors


def write_material_csv(file_path: Path) -> None:
    """写入20条ERP物料CSV。"""
    with file_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["编码", "名称", "规格", "单位", "类别"])
        writer.writerows(SAMPLE_MATERIALS)


async def fake_batch_match(items: list[dict], db, app_state) -> list[match_service.MatchResult]:
    """模拟三级匹配结果，覆盖自动通过、待审核和缺失三种状态。"""
    result_by_name = {
        "铜柱": match_service.MatchResult("铜柱", "MAT-0001", "铜柱", "M3x6", 1.0, "exact", []),
        "主控芯片": match_service.MatchResult(
            "主控芯片",
            "MAT-0002",
            "主控芯片",
            "STM32F103",
            0.92,
            "embedding",
            [{"code": "MAT-0002", "name": "主控芯片", "spec": "STM32F103", "score": 0.92}],
        ),
        "电容100uF": match_service.MatchResult(
            "电容100uF",
            "MAT-0003",
            "电容100uF",
            "贴片",
            0.78,
            "llm",
            [{"code": "MAT-0003", "name": "电容100uF", "spec": "贴片", "score": 0.81}],
        ),
        "未知零件XYZ": match_service.MatchResult("未知零件XYZ", None, None, None, 0.0, "none", []),
    }
    return [result_by_name[item["name"]] for item in items]


async def run_e2e() -> dict:
    """执行完整端到端集成测试。"""
    with tempfile.TemporaryDirectory(prefix="bom_e2e_") as temp_dir:
        temp_path = Path(temp_dir)
        database_path = temp_path / "e2e.db"
        csv_path = temp_path / "sample_materials.csv"
        os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{database_path.as_posix()}"

        engine = create_async_engine(os.environ["DATABASE_URL"], echo=False, future=True)
        session_local = async_sessionmaker(engine, expire_on_commit=False)

        import app.core.database as database_module
        import app.services.material_service as material_service

        old_engine = database_module.engine
        old_session_local = database_module.AsyncSessionLocal
        old_material_session_local = material_service.AsyncSessionLocal
        old_faiss_path = material_service.FAISS_INDEX_PATH
        old_id_map_path = material_service.ID_MAP_PATH
        old_batch_match = match_service.batch_match

        material_service.AsyncSessionLocal = session_local
        material_service.FAISS_INDEX_PATH = temp_path / "materials.faiss"
        material_service.ID_MAP_PATH = temp_path / "id_map.json"
        database_module.engine = engine
        database_module.AsyncSessionLocal = session_local
        match_service.batch_match = fake_batch_match

        try:
            async with engine.begin() as connection:
                await connection.run_sync(Base.metadata.create_all)

            print("Step 1：导入20条示例物料到 materials 表")
            write_material_csv(csv_path)
            import_result = await async_import_materials_from_csv(str(csv_path))
            print(json.dumps(import_result, ensure_ascii=False, indent=2))
            if import_result["success"] != 20:
                raise RuntimeError("示例物料导入数量不正确")

            print("Step 2：构建向量索引 materials.faiss")
            await async_build_embedding_index(embedding_provider=deterministic_embedding)
            index, id_map = load_index()
            if index is None or index.ntotal != 20 or len(id_map) != 20:
                raise RuntimeError("FAISS索引构建失败")
            app_state = SimpleNamespace(material_index=index, material_id_map=id_map)

            print("Step 3：模拟BOM图片识别后的提取结果")
            print(json.dumps(EXTRACTED_BOM, ensure_ascii=False, indent=2))

            print("Step 4：调用匹配引擎处理上述BOM")
            async with session_local() as session:
                stats = await match_service.process_extracted_bom(EXTRACTED_BOM, PRODUCT_NAME, session, app_state)
                print(json.dumps(stats, ensure_ascii=False, indent=2))

                print("Step 5：打印匹配结果，验证置信度分布")
                items_result = await session.execute(select(BomItem).where(BomItem.product_name == PRODUCT_NAME))
                bom_items = list(items_result.scalars().all())
                for item in bom_items:
                    print(
                        f"- {item.raw_name}: status={item.status}, code={item.material_code}, "
                        f"confidence={float(item.confidence or 0):.2f}, level={item.match_level}"
                    )
                missing_result = await session.execute(select(MissingMaterial))
                missing_count = len(list(missing_result.scalars().all()))
                if stats != {"auto_confirmed": 2, "pending_review": 2, "missing": 1, "total": 4}:
                    raise RuntimeError("匹配统计不符合预期")
                if missing_count != 1:
                    raise RuntimeError("缺失物料数量不符合预期")

                print("Step 6：调用导出接口，检查生成的Excel包含4个Sheet")
                excel_bytes = await export_bom_to_excel(PRODUCT_NAME, session)
                excel_path = temp_path / "BOM_主控板V2_e2e.xlsx"
                excel_path.write_bytes(excel_bytes)
                workbook = load_workbook(excel_path)
                sheet_names = workbook.sheetnames
                print(f"Excel Sheets: {sheet_names}")
                if sheet_names != ["BOM导入表", "待处理项", "需新建物料", "操作日志"]:
                    raise RuntimeError("Excel Sheet数量或名称不正确")
                if workbook["BOM导入表"].max_row != 3:
                    raise RuntimeError("BOM导入表应包含2条自动确认记录")

                material_total = int((await session.execute(select(func.count()).select_from(Material))).scalar() or 0)

            return {
                "materials": material_total,
                "index_count": index.ntotal,
                "id_map_count": len(id_map),
                "match_stats": stats,
                "sheets": sheet_names,
            }
        finally:
            match_service.batch_match = old_batch_match
            material_service.AsyncSessionLocal = old_material_session_local
            material_service.FAISS_INDEX_PATH = old_faiss_path
            material_service.ID_MAP_PATH = old_id_map_path
            database_module.engine = old_engine
            database_module.AsyncSessionLocal = old_session_local
            await engine.dispose()


def main() -> None:
    """运行端到端集成测试。"""
    result = asyncio.run(run_e2e())
    print("端到端集成测试通过")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
