import asyncio
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select


PROJECT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = PROJECT_DIR / "backend"
DATABASE_FILE = BACKEND_DIR / "data" / "bom.db"
DATABASE_FILE.parent.mkdir(parents=True, exist_ok=True)
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{DATABASE_FILE.resolve().as_posix()}"
sys.path.insert(0, str(BACKEND_DIR))

from app.core.database import AsyncSessionLocal, init_db  # noqa: E402
from app.models.bom_item import BomItem  # noqa: E402
from app.models.missing_material import MissingMaterial  # noqa: E402
from app.models.operation_log import OperationLog  # noqa: E402
from app.services.export_service import export_bom_to_excel  # noqa: E402


PRODUCT_NAME = "导出测试产品"
OUTPUT_DIR = BACKEND_DIR / "data"


async def clear_old_test_data(session) -> None:
    """清理本脚本上次写入的测试数据。"""
    result = await session.execute(select(BomItem).where(BomItem.product_name == PRODUCT_NAME))
    test_items = list(result.scalars().all())
    item_ids = [item.id for item in test_items]
    if item_ids:
        log_result = await session.execute(select(OperationLog).where(OperationLog.target_id.in_(item_ids)))
        for log in log_result.scalars().all():
            await session.delete(log)
    for item in test_items:
        await session.delete(item)
    missing_result = await session.execute(select(MissingMaterial).where(MissingMaterial.raw_name.like("导出测试%")))
    for missing in missing_result.scalars().all():
        await session.delete(missing)
    await session.commit()


async def insert_test_data(session) -> None:
    """插入混合状态的BOM导出测试数据。"""
    confirmed_items = [
        BomItem(
            product_name=PRODUCT_NAME,
            product_code="EXP-P001",
            material_code=f"EXP-M{i:03d}",
            material_name=name,
            raw_name=name if i <= 3 else f"研发叫法{i}",
            quantity=i,
            unit="个",
            level=1,
            confidence=0.91 + i * 0.001,
            status="confirmed",
            reviewer=None if i <= 3 else "测试审核员",
            match_level="exact" if i <= 3 else "llm",
        )
        for i, name in enumerate(["铜柱", "螺钉", "垫片", "线束", "卡扣", "支架"], start=1)
    ]
    pending_items = [
        BomItem(
            product_name=PRODUCT_NAME,
            product_code="EXP-P001",
            raw_name=f"导出测试待处理{i}",
            quantity=i,
            unit="件",
            level=1,
            confidence=0.55 + i * 0.03,
            status="pending" if i < 4 else "rejected",
            match_level="none" if i < 4 else "llm",
        )
        for i in range(1, 5)
    ]
    session.add_all(confirmed_items + pending_items)
    await session.flush()
    session.add_all(
        [
            MissingMaterial(
                raw_name="导出测试缺失1",
                ai_suggested_name="测试异形件",
                ai_suggested_spec="定制",
                ai_suggested_unit="件",
                ai_suggested_category="加工件",
                status="pending",
            ),
            OperationLog(
                operation="reassign",
                target_id=confirmed_items[-1].id,
                operator="测试审核员",
                before_value='{"material_code": null}',
                after_value='{"material_code": "EXP-M006"}',
            ),
        ]
    )
    await session.commit()


def check_workbook(file_path: Path) -> None:
    """检查导出的Excel内容是否符合基本要求。"""
    workbook = load_workbook(file_path)
    expected_sheets = ["BOM导入表", "待处理项", "需新建物料", "操作日志"]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Sheet不匹配：{workbook.sheetnames}")
    if workbook["BOM导入表"].max_row != 7:
        raise RuntimeError("BOM导入表应包含6条已确认测试数据")
    if workbook["待处理项"].max_row != 5:
        raise RuntimeError("待处理项应包含4条 pending/rejected 测试数据")
    if workbook["需新建物料"].max_row < 2:
        raise RuntimeError("需新建物料应包含至少1条待处理缺失物料")
    if workbook["操作日志"].max_row < 2:
        raise RuntimeError("操作日志应包含测试重分配记录")


async def main() -> None:
    """执行BOM导出测试。"""
    await init_db()
    async with AsyncSessionLocal() as session:
        await clear_old_test_data(session)
        await insert_test_data(session)
        content = await export_bom_to_excel(PRODUCT_NAME, session)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    file_path = OUTPUT_DIR / f"BOM_{PRODUCT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file_path.write_bytes(content)
    check_workbook(file_path)
    print(f"导出测试通过：{file_path}")


if __name__ == "__main__":
    asyncio.run(main())
