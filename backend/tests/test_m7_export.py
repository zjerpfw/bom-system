from io import BytesIO
import sys
from pathlib import Path

import pytest
import pytest_asyncio
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine


BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))


@pytest_asyncio.fixture
async def export_session(tmp_path):
    from app.core.database import Base
    from app.models.bom_item import BomItem
    from app.models.missing_material import MissingMaterial
    from app.models.operation_log import OperationLog

    database_url = f"sqlite+aiosqlite:///{(tmp_path / 'export.db').as_posix()}"
    engine = create_async_engine(database_url, echo=False, future=True)
    session_local = async_sessionmaker(engine, expire_on_commit=False)

    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)

    async with session_local() as session:
        confirmed_auto = BomItem(
            product_name="测试夹具",
            product_code="P001",
            material_code="M001",
            material_name="铜柱",
            raw_name="铜柱",
            quantity=4,
            unit="个",
            level=1,
            confidence=0.95,
            status="confirmed",
            match_level="exact",
        )
        confirmed_manual = BomItem(
            product_name="测试夹具",
            product_code="P001",
            material_code="M002",
            material_name="螺钉",
            raw_name="螺丝",
            quantity=8,
            unit="个",
            level=1,
            confidence=0.88,
            status="confirmed",
            reviewer="张三",
            match_level="llm",
        )
        pending_item = BomItem(
            product_name="测试夹具",
            product_code="P001",
            raw_name="异形件",
            quantity=1,
            unit="件",
            level=1,
            confidence=0.45,
            status="pending",
            match_level="none",
        )
        other_item = BomItem(
            product_name="其他产品",
            product_code="P002",
            material_code="M003",
            material_name="垫片",
            raw_name="垫片",
            quantity=2,
            unit="个",
            level=1,
            confidence=0.96,
            status="confirmed",
        )
        session.add_all([confirmed_auto, confirmed_manual, pending_item, other_item])
        await session.flush()
        session.add_all(
            [
                MissingMaterial(
                    raw_name="异形件",
                    ai_suggested_name="异形件",
                    ai_suggested_spec="定制",
                    ai_suggested_unit="件",
                    ai_suggested_category="加工件",
                    status="pending",
                ),
                OperationLog(
                    operation="reassign",
                    target_id=confirmed_manual.id,
                    operator="张三",
                    before_value='{"material_code": null}',
                    after_value='{"material_code": "M002"}',
                ),
            ]
        )
        await session.commit()
        yield session

    await engine.dispose()


@pytest.mark.asyncio
async def test_export_bom_to_excel_creates_four_sheets(export_session):
    from app.services.export_service import export_bom_to_excel

    content = await export_bom_to_excel("测试夹具", export_session)
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["BOM导入表", "待处理项", "需新建物料", "操作日志"]

    bom_sheet = workbook["BOM导入表"]
    pending_sheet = workbook["待处理项"]
    missing_sheet = workbook["需新建物料"]
    log_sheet = workbook["操作日志"]

    assert [cell.value for cell in bom_sheet[1]] == ["父件编码", "父件名称", "子件编码", "子件名称", "规格", "用量", "单位", "层级"]
    assert bom_sheet.max_row == 3
    assert bom_sheet["A2"].value == "P001"
    assert bom_sheet["C2"].value == "M001"
    assert bom_sheet["F2"].value == 4
    assert bom_sheet["A1"].fill.fgColor.rgb == "001D6FA5"
    assert bom_sheet["A2"].fill.fgColor.rgb == "FFE2F3EA"
    assert bom_sheet["A3"].fill.fgColor.rgb == "00FFFFFF"
    assert bom_sheet.column_dimensions["A"].width >= 10

    assert pending_sheet.max_row == 2
    assert pending_sheet["A2"].value == "异形件"
    assert pending_sheet["A2"].font.color.rgb == "00D64545"

    assert missing_sheet.max_row == 2
    assert missing_sheet["A2"].value == "异形件"
    assert log_sheet.max_row == 2
    assert log_sheet["B2"].value == "reassign"


def test_export_routes_return_excel(tmp_path, monkeypatch):
    from app.core.database import Base, get_db
    from app.models.bom_item import BomItem
    import main
    import asyncio

    database_url = f"sqlite+aiosqlite:///{(tmp_path / 'export_api.db').as_posix()}"
    engine = create_async_engine(database_url, echo=False, future=True)
    session_local = async_sessionmaker(engine, expire_on_commit=False)

    async def create_tables():
        async with engine.begin() as connection:
            await connection.run_sync(Base.metadata.create_all)
        async with session_local() as session:
            session.add(
                BomItem(
                    product_name="测试夹具",
                    product_code="P001",
                    material_code="M001",
                    material_name="铜柱",
                    raw_name="铜柱",
                    quantity=4,
                    unit="个",
                    level=1,
                    confidence=0.95,
                    status="confirmed",
                )
            )
            await session.commit()

    async def override_get_db():
        async with session_local() as session:
            yield session

    asyncio.run(create_tables())
    main.app.dependency_overrides[get_db] = override_get_db
    client = TestClient(main.app)

    response = client.get("/api/export/bom/测试夹具")
    pending_response = client.get("/api/export/all-pending")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")
    assert "BOM_" in response.headers["content-disposition"]
    assert load_workbook(BytesIO(response.content)).sheetnames == ["BOM导入表", "待处理项", "需新建物料", "操作日志"]
    assert pending_response.status_code == 200
    assert load_workbook(BytesIO(pending_response.content)).sheetnames == ["BOM导入表", "待处理项", "需新建物料", "操作日志"]
    main.app.dependency_overrides.clear()
