from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bom_item import BomItem
from app.models.missing_material import MissingMaterial
from app.models.operation_log import OperationLog


HEADER_FILL = PatternFill("solid", fgColor="1D6FA5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
AUTO_CONFIRM_FILL = PatternFill("solid", fgColor="FFE2F3EA")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
ERROR_FONT = Font(color="D64545")
NUMBER_ALIGNMENT = Alignment(horizontal="right")


def format_cell_value(value):
    """格式化Excel单元格值。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, Decimal):
        return float(value) if value != int(value) else int(value)
    if isinstance(value, float):
        return value if value != int(value) else int(value)
    return value


def apply_header_style(sheet) -> None:
    """设置表头样式。"""
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def auto_fit_columns(sheet) -> None:
    """按内容自适应列宽。"""
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def align_number_columns(sheet, columns: set[int]) -> None:
    """设置数字列右对齐。"""
    for row in sheet.iter_rows(min_row=2):
        for index in columns:
            row[index - 1].alignment = NUMBER_ALIGNMENT


def append_header(sheet, headers: list[str]) -> None:
    """写入表头。"""
    sheet.append(headers)
    apply_header_style(sheet)


async def get_bom_items(product_name: str | None, db: AsyncSession) -> list[BomItem]:
    """查询BOM条目。"""
    statement = select(BomItem)
    if product_name:
        statement = statement.where(BomItem.product_name == product_name)
    result = await db.execute(statement.order_by(BomItem.id))
    return list(result.scalars().all())


async def get_missing_materials(db: AsyncSession) -> list[MissingMaterial]:
    """查询待新建物料。"""
    result = await db.execute(
        select(MissingMaterial).where(MissingMaterial.status == "pending").order_by(MissingMaterial.id)
    )
    return list(result.scalars().all())


async def get_operation_logs(product_name: str | None, db: AsyncSession) -> list[OperationLog]:
    """查询产品相关操作日志。"""
    if not product_name:
        result = await db.execute(select(OperationLog).order_by(OperationLog.id))
        return list(result.scalars().all())

    item_result = await db.execute(select(BomItem.id).where(BomItem.product_name == product_name))
    item_ids = [row[0] for row in item_result.all()]
    if not item_ids:
        return []
    log_result = await db.execute(select(OperationLog).where(OperationLog.target_id.in_(item_ids)).order_by(OperationLog.id))
    return list(log_result.scalars().all())


def fill_bom_sheet(sheet, items: list[BomItem]) -> None:
    """填充BOM导入表。"""
    append_header(sheet, ["父件编码", "父件名称", "子件编码", "子件名称", "规格", "用量", "单位", "层级"])
    for item in items:
        if item.status != "confirmed":
            continue
        sheet.append(
            [
                item.product_code,
                item.product_name,
                item.material_code,
                item.material_name,
                "",
                format_cell_value(item.quantity),
                item.unit,
                item.level,
            ]
        )
        fill = AUTO_CONFIRM_FILL if not item.reviewer else WHITE_FILL
        for cell in sheet[sheet.max_row]:
            cell.fill = fill
    align_number_columns(sheet, {6, 8})


def fill_pending_sheet(sheet, items: list[BomItem]) -> None:
    """填充待处理项。"""
    append_header(sheet, ["原始叫法", "匹配状态", "置信度", "备注"])
    for item in items:
        if item.status not in {"pending", "rejected"}:
            continue
        sheet.append([item.raw_name, item.status, float(item.confidence or 0), item.match_level or ""])
        for cell in sheet[sheet.max_row]:
            cell.font = ERROR_FONT
    align_number_columns(sheet, {3})


def fill_missing_sheet(sheet, missing_materials: list[MissingMaterial]) -> None:
    """填充需新建物料。"""
    append_header(sheet, ["建议名称", "建议规格", "建议单位", "建议类别", "状态"])
    for item in missing_materials:
        sheet.append(
            [
                item.ai_suggested_name or item.raw_name,
                item.ai_suggested_spec,
                item.ai_suggested_unit,
                item.ai_suggested_category,
                item.status,
            ]
        )


def fill_log_sheet(sheet, logs: list[OperationLog]) -> None:
    """填充操作日志。"""
    append_header(sheet, ["时间", "操作", "原始值", "修改值", "操作人"])
    for log in logs:
        sheet.append(
            [
                format_cell_value(log.created_at),
                log.operation,
                log.before_value,
                log.after_value,
                log.operator,
            ]
        )


async def export_bom_to_excel(product_name: str, db: AsyncSession) -> bytes:
    """导出某产品BOM导入包。"""
    return await build_export_workbook(product_name, db)


async def export_all_pending_to_excel(db: AsyncSession) -> bytes:
    """导出所有待处理项汇总。"""
    return await build_export_workbook(None, db)


async def build_export_workbook(product_name: str | None, db: AsyncSession) -> bytes:
    """构建Excel导出文件。"""
    items = await get_bom_items(product_name, db)
    missing_materials = await get_missing_materials(db)
    logs = await get_operation_logs(product_name, db)

    workbook = Workbook()
    bom_sheet = workbook.active
    bom_sheet.title = "BOM导入表"
    pending_sheet = workbook.create_sheet("待处理项")
    missing_sheet = workbook.create_sheet("需新建物料")
    log_sheet = workbook.create_sheet("操作日志")

    fill_bom_sheet(bom_sheet, items)
    fill_pending_sheet(pending_sheet, items)
    fill_missing_sheet(missing_sheet, missing_materials)
    fill_log_sheet(log_sheet, logs)

    for sheet in workbook.worksheets:
        auto_fit_columns(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
